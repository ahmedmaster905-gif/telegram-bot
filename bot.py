import csv
import os
import asyncio
import shutil
import zipfile
import logging
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackQueryHandler, filters, ContextTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==================== إعدادات التسجيل ====================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ==================== الإعدادات الأساسية ====================

TOKEN = "8751952643:AAEvkGQYrc_teXAxedurciQQKRrt4KBKEos"
GROUP_ID = -1003621204878
TEACHERS_FILE = "teachers.txt"
SUPERVISORS_FILE = "supervisors.txt"
REPORTS_FOLDER = "reports"
ARCHIVE_FOLDER = "archive"
BACKUP_FOLDER = "backups"

# ==================== المتغيرات العامة ====================

questions = {}
teacher_state = {}
student_active_session = {}
student_messages_in_group = {}
teacher_ratings = {}
pending_ratings = {}
monthly_stats = {}
teacher_active_questions = {}

teacher_ids = []
supervisor_ids = []

DEVELOPER_ID = 8420041691
bot_is_admin = False
current_month = datetime.now().strftime("%Y-%m")
month_active = True
MAX_ACTIVE_QUESTIONS = 1

for folder in [REPORTS_FOLDER, ARCHIVE_FOLDER, BACKUP_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# ==================== دوال إدارة الملفات ====================

def load_teachers():
    global teacher_ids
    if os.path.exists(TEACHERS_FILE):
        with open(TEACHERS_FILE, 'r', encoding='utf-8') as f:
            teacher_ids = [int(line.strip()) for line in f if line.strip()]
    else:
        teacher_ids = []

def save_teachers():
    with open(TEACHERS_FILE, 'w', encoding='utf-8') as f:
        for tid in teacher_ids:
            f.write(f"{tid}\n")

def load_supervisors():
    global supervisor_ids
    if os.path.exists(SUPERVISORS_FILE):
        with open(SUPERVISORS_FILE, 'r', encoding='utf-8') as f:
            supervisor_ids = [int(line.strip()) for line in f if line.strip()]
    else:
        supervisor_ids = []

def save_supervisors():
    with open(SUPERVISORS_FILE, 'w', encoding='utf-8') as f:
        for sid in supervisor_ids:
            f.write(f"{sid}\n")

def load_ratings():
    global teacher_ratings
    ratings_file = "ratings.csv"
    if os.path.exists(ratings_file):
        with open(ratings_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 3:
                    tid = int(row[0])
                    teacher_ratings[tid] = {
                        "total_rating": float(row[1]),
                        "count": int(row[2])
                    }

def save_ratings():
    ratings_file = "ratings.csv"
    with open(ratings_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Teacher ID", "Total Rating", "Rating Count", "Average"])
        for tid, data in teacher_ratings.items():
            avg = data["total_rating"] / data["count"] if data["count"] > 0 else 0
            writer.writerow([tid, data["total_rating"], data["count"], round(avg, 2)])

def load_questions():
    global questions
    filename = f"questions_{current_month}.csv"
    if os.path.exists(filename):
        with open(filename, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 14:
                    qid = row[0]
                    group_msg_id = None
                    if len(row) > 13 and row[13] and row[13] != "" and row[13] != "False" and row[13] != "True":
                        try:
                            group_msg_id = int(row[13])
                        except ValueError:
                            group_msg_id = None
                    
                    questions[qid] = {
                        "student_id": int(row[1]),
                        "student_name": row[2],
                        "teacher_id": int(row[3]) if row[3] and row[3] != "" else None,
                        "status": row[4],
                        "count_for_stats": row[5].lower() == "true",
                        "rating": row[6] if row[6] else "",
                        "text": row[7],
                        "file_id": row[8] if row[8] else None,
                        "file_type": row[9] if row[9] else None,
                        "file_name": row[10] if row[10] else None,
                        "reply": row[11] if row[11] else "",
                        "reply_time": row[12] if row[12] else "",
                        "group_message_id": group_msg_id
                    }

def save_questions():
    filename = f"questions_{current_month}.csv"
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["QID", "Student ID", "Student Name", "Teacher ID", "Status", "Counted", "Rating", "Text", "File ID", "File Type", "File Name", "Reply", "Reply Time", "Group Message ID"])
        for qid, q in questions.items():
            writer.writerow([
                qid, q["student_id"], q["student_name"], q.get("teacher_id", ""),
                q["status"], q["count_for_stats"], q.get("rating", ""), q["text"],
                q.get("file_id", ""), q.get("file_type", ""), q.get("file_name", ""),
                q.get("reply", ""), q.get("reply_time", ""), q.get("group_message_id", "")
            ])

def load_monthly_stats():
    global monthly_stats, current_month, month_active
    stats_file = f"monthly_stats_{current_month}.csv"
    if os.path.exists(stats_file):
        with open(stats_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if len(row) >= 5:
                    tid = int(row[0])
                    monthly_stats[tid] = {
                        "total": int(row[1]),
                        "counted": int(row[2]),
                        "rating_sum": float(row[3]),
                        "rating_count": int(row[4])
                    }
    status_file = "month_status.txt"
    if os.path.exists(status_file):
        with open(status_file, 'r', encoding='utf-8') as f:
            data = f.read().strip().split(",")
            if len(data) >= 2:
                current_month = data[0]
                month_active = data[1] == "active"

def save_monthly_stats():
    stats_file = f"monthly_stats_{current_month}.csv"
    with open(stats_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Teacher ID", "Total", "Counted", "Rating Sum", "Rating Count"])
        for tid, data in monthly_stats.items():
            writer.writerow([tid, data["total"], data["counted"], data["rating_sum"], data["rating_count"]])
    
    with open("month_status.txt", 'w', encoding='utf-8') as f:
        f.write(f"{current_month},{'active' if month_active else 'closed'}")

def update_monthly_stats(teacher_id, is_counted, rating=0):
    if not month_active:
        return
    if teacher_id not in monthly_stats:
        monthly_stats[teacher_id] = {"total": 0, "counted": 0, "rating_sum": 0, "rating_count": 0}
    monthly_stats[teacher_id]["total"] += 1
    if is_counted:
        monthly_stats[teacher_id]["counted"] += 1
    if rating > 0:
        monthly_stats[teacher_id]["rating_sum"] += rating
        monthly_stats[teacher_id]["rating_count"] += 1
    save_monthly_stats()

def update_teacher_load(teacher_id, change):
    if teacher_id not in teacher_active_questions:
        teacher_active_questions[teacher_id] = 0
    teacher_active_questions[teacher_id] += change
    if teacher_active_questions[teacher_id] < 0:
        teacher_active_questions[teacher_id] = 0
    return teacher_active_questions[teacher_id]

# ==================== دوال النسخ الاحتياطي ====================

async def create_backup():
    backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    backup_path = os.path.join(BACKUP_FOLDER, backup_name)
    
    with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        files = [TEACHERS_FILE, SUPERVISORS_FILE, "ratings.csv", f"monthly_stats_{current_month}.csv", "month_status.txt", f"questions_{current_month}.csv"]
        for f in files:
            if os.path.exists(f):
                zipf.write(f, os.path.basename(f))
    
    backups = sorted([f for f in os.listdir(BACKUP_FOLDER) if f.startswith("backup_")])
    while len(backups) > 7:
        os.remove(os.path.join(BACKUP_FOLDER, backups.pop(0)))
    
    return backup_path

# ==================== دوال البوت الأساسية ====================

async def check_admin_status(bot, group_id):
    global bot_is_admin
    try:
        bot_member = await bot.get_chat_member(chat_id=group_id, user_id=bot.id)
        if bot_member.status in ['administrator', 'creator']:
            bot_is_admin = True
            logger.info("✅ البوت أدمن في الجروب")
        else:
            bot_is_admin = False
            logger.info("❌ البوت ليس أدمن")
        return bot_is_admin
    except Exception as e:
        logger.error(f"⚠️ فشل التحقق: {e}")
        try:
            test_msg = await bot.send_message(chat_id=group_id, text="🔍 اختبار صلاحية البوت...")
            await bot.delete_message(chat_id=group_id, message_id=test_msg.message_id)
            bot_is_admin = True
            logger.info("✅ البوت يستطيع حذف الرسائل")
        except:
            bot_is_admin = False
        return bot_is_admin

async def delete_student_messages(context, student_id, group_id):
    if not bot_is_admin or student_id not in student_messages_in_group:
        return 0
    deleted = 0
    for msg in student_messages_in_group[student_id].copy():
        try:
            await context.bot.delete_message(group_id, msg.get("message_id"))
            deleted += 1
        except:
            pass
    if student_id in student_messages_in_group:
        del student_messages_in_group[student_id]
    return deleted

async def remove_buttons_from_group_message(context, group_id, message_id):
    """إزالة الأزرار من رسالة في الجروب"""
    if not message_id:
        return False
    try:
        await context.bot.edit_message_reply_markup(chat_id=group_id, message_id=message_id, reply_markup=None)
        return True
    except Exception as e:
        logger.error(f"فشل إزالة الأزرار: {e}")
        return False

async def send_question_to_teacher(context, teacher_id, qid, q, is_converted=False):
    file_id = q.get("file_id")
    text = q["text"]
    student_name = q["student_name"]
    student_code = q["student_id"]
    file_name = q.get("file_name", "")
    is_counted = q.get("count_for_stats", True)
    counted_status = "✅ محسوب" if is_counted else "❌ غير محسوب"
    
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🚫 لا يتحسب", callback_data=f"skip|{qid}")],
        [InlineKeyboardButton("🔄 إعادة للجروب", callback_data=f"return|{qid}")],
        [InlineKeyboardButton("🔚 إنهاء الجلسة", callback_data=f"end|{qid}")]
    ])
    
    msg = f"📬 **{'سؤال محول' if is_converted else 'سؤال جديد'} من {student_name}**\n"
    msg += f"🔢 كود: {student_code}\n📊 الحالة: {counted_status}\n\n📝 {text}\n\n---\n"
    msg += "📌 الأزرار:\n• 🚫 لا يتحسب - السؤال لا يدخل في الإحصاء\n• 🔄 إعادة للجروب - إعادة السؤال للجروب لمدرس آخر\n• 🔚 إنهاء الجلسة - إنهاء الجلسة وتقييم الطالب"
    
    try:
        if file_id:
            if q["file_type"] == "photo":
                await context.bot.send_photo(teacher_id, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif q["file_type"] == "video":
                await context.bot.send_video(teacher_id, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif q["file_type"] == "audio":
                await context.bot.send_audio(teacher_id, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif q["file_type"] == "voice":
                await context.bot.send_voice(teacher_id, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif q["file_type"] == "document":
                await context.bot.send_document(teacher_id, file_id, filename=file_name, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
        else:
            await context.bot.send_message(teacher_id, msg, reply_markup=keyboard, parse_mode="Markdown")
        return True
    except Exception as e:
        logger.error(f"Send to teacher error: {e}")
        return False

async def send_all_questions_to_teacher(context, student_id, teacher_id):
    sent = 0
    for qid, q in questions.items():
        if q["student_id"] == student_id and q.get("teacher_id") is None:
            q["teacher_id"] = teacher_id
            q["status"] = "taken"
            if await send_question_to_teacher(context, teacher_id, qid, q, True):
                sent += 1
                await asyncio.sleep(0.3)
    return sent

async def return_to_group(context, teacher_id, qid):
    """إعادة السؤال للجروب"""
    if qid not in questions:
        return False
    q = questions[qid]
    student_id = q["student_id"]
    
    q["teacher_id"] = None
    q["status"] = "open"
    
    if teacher_id in teacher_state:
        del teacher_state[teacher_id]
    if student_id in student_active_session:
        del student_active_session[student_id]
    update_teacher_load(teacher_id, -1)
    save_questions()
    
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("📥 استلام السؤال", callback_data=f"take|{qid}|{student_id}")]])
    
    if q.get("group_message_id"):
        try:
            await context.bot.edit_message_reply_markup(
                chat_id=GROUP_ID,
                message_id=q["group_message_id"],
                reply_markup=keyboard
            )
        except Exception as e:
            logger.error(f"فشل تحديث رسالة الجروب: {e}")
    
    await context.bot.send_message(teacher_id, "✅ تم إعادة السؤال للجروب")
    return True

async def send_end_session_message(context, student_id, teacher_name):
    """إرسال رسالة للطالب بانتهاء الجلسة"""
    try:
        await context.bot.send_message(
            student_id,
            f"🔒 **انتهت الجلسة**\n\n"
            f"👨‍🏫 المدرس {teacher_name} أنهى الجلسة.\n"
            f"📝 سيظهر لك الآن طلب تقييم للمدرس.\n\n"
            f"⭐ الرجاء تقييم المدرس من 1 إلى 5 نجوم.",
            parse_mode="Markdown"
        )
    except Exception as e:
        logger.error(f"Failed to send end session message: {e}")

def create_excel_report(period_name):
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86AB")
    border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    
    ws = wb.active
    ws.title = "إحصائيات المدرسين"
    headers = ["رقم المدرس", "عدد الأسئلة", "المحسوب", "غير المحسوب", "متوسط التقييم"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font, cell.fill, cell.border, cell.alignment = header_font, header_fill, border, Alignment(horizontal="center")
    
    stats = {}
    for q in questions.values():
        tid = q.get("teacher_id")
        if tid:
            stats.setdefault(tid, {"total": 0, "counted": 0})
            stats[tid]["total"] += 1
            if q.get("count_for_stats"):
                stats[tid]["counted"] += 1
    
    for i, (tid, s) in enumerate(stats.items(), 2):
        rating = teacher_ratings.get(tid, {"total_rating": 0, "count": 0})
        avg = rating["total_rating"] / rating["count"] if rating["count"] else 0
        ws.cell(i, 1, tid).border = border
        ws.cell(i, 2, s["total"]).border = border
        ws.cell(i, 3, s["counted"]).border = border
        ws.cell(i, 4, s["total"] - s["counted"]).border = border
        ws.cell(i, 5, round(avg, 2)).border = border
    
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    filename = f"{REPORTS_FOLDER}/report_{period_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    return filename

# ==================== الأوامر ====================

async def start(update: Update, context):
    uid = update.effective_user.id
    status = "🟢 مفتوح" if month_active else "🔴 مغلق"
    msg = f"🎓 **مرحباً بك!**\n\n📤 أرسل سؤالك (نص، صورة، فيديو، صوت، ملف)\n❌ /cancel - إلغاء سؤال قبل استلامه\n📅 حالة الشهر: {status}\n📌 /help - عرض الأوامر"
    await update.message.reply_text(msg, parse_mode="Markdown")

async def help_command(update: Update, context):
    uid = update.effective_user.id
    text = "📚 **الأوامر**\n\n👥 **للكل:** /start, /myid, /help\n🎓 **للطلاب:** /cancel\n👨‍🏫 **للمدرسين:** /mystats, /myrating, /report, /mystatus\n👁️ **للمراقبين:** /allreplies, /teacherstats, /teachers, /addteacher, /monthstatus\n👨‍💻 **للمطور:** /removeteacher, /addsupervisor, /removesupervisor, /listsupervisors, /checkadmin, /startmonth, /closemonth, /backup"
    await update.message.reply_text(text, parse_mode="Markdown")

async def myid(update: Update, context):
    uid = update.effective_user.id
    role = "✅ مدرس" if uid in teacher_ids else "👁️ مراقب" if uid in supervisor_ids else "👨‍💻 مطور" if uid == DEVELOPER_ID else "❌ طالب"
    await update.message.reply_text(f"🆔 **رقمك:** `{uid}`\n📚 **الصلاحية:** {role}", parse_mode="Markdown")

async def mystats(update: Update, context):
    uid = update.effective_user.id
    if uid not in teacher_ids:
        return await update.message.reply_text("⚠️ للمدرسين فقط")
    total = sum(1 for q in questions.values() if q.get("teacher_id") == uid)
    counted = sum(1 for q in questions.values() if q.get("teacher_id") == uid and q.get("count_for_stats"))
    rating = teacher_ratings.get(uid, {"total_rating": 0, "count": 0})
    avg = rating["total_rating"] / rating["count"] if rating["count"] else 0
    monthly = monthly_stats.get(uid, {"total": 0, "counted": 0})
    await update.message.reply_text(f"📊 **إحصائياتك**\n\n✅ المحسوب: {counted}\n📦 الإجمالي: {total}\n⭐ التقييم: {round(avg, 2)}/5\n📊 التقييمات: {rating['count']}\n📅 الشهر: {monthly['total']} سؤال, {monthly['counted']} محسوب", parse_mode="Markdown")

async def myrating(update: Update, context):
    uid = update.effective_user.id
    if uid not in teacher_ids:
        return await update.message.reply_text("⚠️ للمدرسين فقط")
    data = teacher_ratings.get(uid, {"total_rating": 0, "count": 0})
    avg = data["total_rating"] / data["count"] if data["count"] else 0
    stars = "⭐" * int(round(avg)) + "☆" * (5 - int(round(avg)))
    await update.message.reply_text(f"📊 **تقييماتك**\n\n{stars}\n📈 المتوسط: {round(avg, 2)}/5\n📊 التقييمات: {data['count']}", parse_mode="Markdown")

async def mystatus(update: Update, context):
    teacher_id = update.effective_user.id
    if teacher_id not in teacher_ids:
        return await update.message.reply_text("⚠️ هذا الأمر للمدرسين فقط.")
    
    if teacher_id in teacher_state:
        qid = teacher_state[teacher_id]["question_id"]
        await update.message.reply_text(
            f"📊 **حالتك الحالية**\n\n"
            f"✅ لديك سؤال نشط\n"
            f"👨‍🎓 الطالب: {questions[qid]['student_name']}\n\n"
            f"🔚 لإنهاء الجلسة، اضغط على زر 'إنهاء الجلسة'\n"
            f"🔄 لإعادة السؤال للجروب، اضغط على زر 'إعادة للجروب'"
        )
    else:
        await update.message.reply_text(
            f"📊 **حالتك الحالية**\n\n"
            f"✅ متاح لاستلام أسئلة جديدة\n"
            f"📥 استلم سؤال من الجروب"
        )

async def report(update: Update, context):
    stats = {}
    for q in questions.values():
        tid = q.get("teacher_id")
        if tid and q.get("count_for_stats"):
            stats[tid] = stats.get(tid, 0) + 1
    if not stats:
        return await update.message.reply_text("📊 لا توجد إحصائيات")
    text = "📊 **تقرير المدرسين**\n\n" + "\n".join(f"👨‍🏫 {tid}: {c} سؤال" for tid, c in stats.items())
    await update.message.reply_text(text, parse_mode="Markdown")

async def cancel_question(update: Update, context):
    uid = update.effective_user.id
    open_q = [(qid, q) for qid, q in questions.items() if q["student_id"] == uid and q["status"] == "open"]
    if not open_q:
        return await update.message.reply_text("❌ لا يوجد أسئلة مفتوحة")
    
    if len(open_q) == 1:
        qid, q = open_q[0]
        if q.get("group_message_id"):
            try:
                await context.bot.delete_message(GROUP_ID, q["group_message_id"])
            except:
                pass
        del questions[qid]
        save_questions()
        if uid in student_messages_in_group:
            for msg in student_messages_in_group[uid]:
                try:
                    await context.bot.delete_message(GROUP_ID, msg.get("message_id"))
                except:
                    pass
            del student_messages_in_group[uid]
        await update.message.reply_text(f"✅ تم إلغاء سؤالك:\n{q['text']}")
    else:
        keyboard = [[InlineKeyboardButton(q['text'][:30], callback_data=f"cancel|{qid}")] for qid, q in open_q[:5]]
        keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="cancel_none")])
        await update.message.reply_text("📋 اختر السؤال للإلغاء:", reply_markup=InlineKeyboardMarkup(keyboard))

async def backup_command(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    await update.message.reply_text("📦 جاري إنشاء النسخة...")
    path = await create_backup()
    with open(path, 'rb') as f:
        await update.message.reply_document(f, filename=os.path.basename(path))

async def start_month(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    global current_month, month_active, monthly_stats
    if month_active:
        return await update.message.reply_text("⚠️ يوجد شهر مفتوح. أغلقه أولاً بـ /closemonth")
    current_month = datetime.now().strftime("%Y-%m")
    monthly_stats = {}
    month_active = True
    save_monthly_stats()
    await update.message.reply_text(f"✅ تم بدء شهر {current_month}")

async def close_month(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    
    global month_active, current_month, monthly_stats, questions, teacher_ratings, teacher_state, student_active_session, teacher_active_questions, pending_ratings, student_messages_in_group
    
    if not month_active:
        return await update.message.reply_text("⚠️ لا يوجد شهر مفتوح")
    
    await update.message.reply_text("📊 جاري إنشاء التقرير النهائي للشهر...")
    
    month_active = False
    month_name = current_month
    save_monthly_stats()
    
    filename = create_excel_report(month_name)
    with open(filename, 'rb') as f:
        await context.bot.send_document(DEVELOPER_ID, f, filename=f"التقرير_النهائي_{month_name}.xlsx")
    
    for sid in supervisor_ids:
        try:
            with open(filename, 'rb') as f:
                await context.bot.send_document(sid, f, filename=f"التقرير_النهائي_{month_name}.xlsx")
            await asyncio.sleep(0.5)
        except:
            pass
    
    await update.message.reply_text(f"✅ **تم إغلاق شهر {month_name}**\n📊 تم إرسال التقرير النهائي للمطور والمراقبين\n\n🔄 جاري تصفير البيانات للشهر الجديد...")
    
    monthly_stats = {}
    questions = {}
    teacher_ratings = {}
    teacher_state = {}
    student_active_session = {}
    teacher_active_questions = {}
    pending_ratings = {}
    student_messages_in_group = {}
    
    save_monthly_stats()
    save_questions()
    save_ratings()
    
    current_month = datetime.now().strftime("%Y-%m")
    month_active = True
    save_monthly_stats()
    
    await update.message.reply_text(
        f"✅ **تم تصفير جميع البيانات**\n\n"
        f"📅 **الشهر الجديد:** {current_month}\n"
        f"🔓 **الحالة:** مفتوح للحساب\n\n"
        f"🚀 يمكن للمدرسين بدء استلام الأسئلة"
    )

async def month_status(update: Update, context):
    uid = update.effective_user.id
    if uid not in supervisor_ids and uid != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمراقبين فقط")
    text = f"📅 **الشهر:** {current_month}\n🔓 **الحالة:** {'🟢 مفتوح' if month_active else '🔴 مغلق'}"
    if month_active and monthly_stats:
        total = sum(d["total"] for d in monthly_stats.values())
        counted = sum(d["counted"] for d in monthly_stats.values())
        text += f"\n📊 الإجمالي: {total}\n✅ المحسوب: {counted}\n📈 النسبة: {round(counted/total*100, 1) if total else 0}%"
    await update.message.reply_text(text, parse_mode="Markdown")

async def check_admin(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    
    await update.message.reply_text("🔍 جاري فحص صلاحيات البوت...")
    
    try:
        bot_info = await context.bot.get_me()
        chat = await context.bot.get_chat(GROUP_ID)
        bot_member = await context.bot.get_chat_member(GROUP_ID, bot_info.id)
        
        result = f"📊 **تقرير صلاحيات البوت**\n\n"
        result += f"👤 **اسم البوت:** {bot_info.first_name}\n"
        result += f"🆔 **ID البوت:** {bot_info.id}\n"
        result += f"👥 **اسم الجروب:** {chat.title}\n\n"
        
        if bot_member.status == 'creator':
            result += "✅ البوت هو منشئ الجروب\n"
        elif bot_member.status == 'administrator':
            result += "✅ البوت أدمن\n"
            result += f"   • حذف الرسائل: {'✅' if bot_member.can_delete_messages else '❌'}\n"
        else:
            result += f"❌ {bot_member.status}\n"
        
        try:
            test_msg = await context.bot.send_message(GROUP_ID, "🧪 اختبار حذف رسالة")
            await context.bot.delete_message(GROUP_ID, test_msg.message_id)
            result += f"\n🧪 **اختبار الحذف:** ✅ نجح\n"
        except Exception as e:
            result += f"\n🧪 **اختبار الحذف:** ❌ فشل\n"
        
        await update.message.reply_text(result, parse_mode='Markdown')
    except Exception as e:
        await update.message.reply_text(f"❌ فشل الفحص: {str(e)}")

# ==================== دوال إدارة المدرسين والمراقبين ====================

async def teachers_list(update: Update, context):
    uid = update.effective_user.id
    if uid not in supervisor_ids and uid != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمراقبين فقط")
    if not teacher_ids:
        return await update.message.reply_text("📭 لا يوجد مدرسون")
    text = "📋 **المدرسون:**\n" + "\n".join(f"{i+1}. {tid}" for i, tid in enumerate(teacher_ids))
    await update.message.reply_text(text, parse_mode="Markdown")

async def add_teacher(update: Update, context):
    uid = update.effective_user.id
    if uid not in supervisor_ids and uid != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمراقبين والمطور فقط")
    if not context.args:
        return await update.message.reply_text("⚠️ /addteacher رقم")
    try:
        tid = int(context.args[0])
        if tid not in teacher_ids:
            teacher_ids.append(tid)
            save_teachers()
            await update.message.reply_text(f"✅ تم إضافة المدرس {tid}")
        else:
            await update.message.reply_text("⚠️ موجود مسبقاً")
    except:
        await update.message.reply_text("⚠️ رقم غير صحيح")

async def remove_teacher(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    if not context.args:
        return await update.message.reply_text("⚠️ /removeteacher رقم")
    try:
        tid = int(context.args[0])
        if tid in teacher_ids:
            teacher_ids.remove(tid)
            save_teachers()
            await update.message.reply_text(f"✅ تم حذف المدرس {tid}")
        else:
            await update.message.reply_text("⚠️ غير موجود")
    except:
        await update.message.reply_text("⚠️ رقم غير صحيح")

async def add_supervisor(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    if not context.args:
        return await update.message.reply_text("⚠️ /addsupervisor رقم")
    try:
        sid = int(context.args[0])
        if sid not in supervisor_ids:
            supervisor_ids.append(sid)
            save_supervisors()
            await update.message.reply_text(f"✅ تم إضافة المراقب {sid}")
        else:
            await update.message.reply_text("⚠️ موجود مسبقاً")
    except:
        await update.message.reply_text("⚠️ رقم غير صحيح")

async def remove_supervisor(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    if not context.args:
        return await update.message.reply_text("⚠️ /removesupervisor رقم")
    try:
        sid = int(context.args[0])
        if sid in supervisor_ids:
            supervisor_ids.remove(sid)
            save_supervisors()
            await update.message.reply_text(f"✅ تم حذف المراقب {sid}")
        else:
            await update.message.reply_text("⚠️ غير موجود")
    except:
        await update.message.reply_text("⚠️ رقم غير صحيح")

async def list_supervisors(update: Update, context):
    if update.effective_user.id != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمطور فقط")
    if not supervisor_ids:
        return await update.message.reply_text("📭 لا يوجد مراقبون")
    text = "👁️ **المراقبون:**\n" + "\n".join(f"{i+1}. {sid}" for i, sid in enumerate(supervisor_ids))
    await update.message.reply_text(text, parse_mode="Markdown")

async def all_replies(update: Update, context):
    uid = update.effective_user.id
    if uid not in supervisor_ids and uid != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمراقبين فقط")
    replied = [(qid, q) for qid, q in questions.items() if q.get("reply")]
    if not replied:
        return await update.message.reply_text("📭 لا يوجد ردود")
    
    for i, (qid, q) in enumerate(replied[-5:][::-1], 1):
        stars = "⭐" * int(round(q.get("rating", 0))) if q.get("rating") else "لا يوجد"
        msg = f"📋 **الرد #{i}**\n👨‍🎓 {q['student_name']}\n👨‍🏫 {q.get('teacher_id')}\n⭐ {stars}\n📝 {q['text'][:100]}\n💬 {q['reply'][:100]}"
        await update.message.reply_text(msg, parse_mode="Markdown")

async def teacher_stats(update: Update, context):
    uid = update.effective_user.id
    if uid not in supervisor_ids and uid != DEVELOPER_ID:
        return await update.message.reply_text("⚠️ للمراقبين فقط")
    if not context.args:
        return await update.message.reply_text("⚠️ /teacherstats رقم")
    try:
        tid = int(context.args[0])
        total = sum(1 for q in questions.values() if q.get("teacher_id") == tid)
        counted = sum(1 for q in questions.values() if q.get("teacher_id") == tid and q.get("count_for_stats"))
        rating = teacher_ratings.get(tid, {"total_rating": 0, "count": 0})
        avg = rating["total_rating"] / rating["count"] if rating["count"] else 0
        monthly = monthly_stats.get(tid, {"total": 0, "counted": 0})
        await update.message.reply_text(f"📊 **المدرس {tid}**\n📦 الإجمالي: {total}\n✅ المحسوب: {counted}\n⭐ التقييم: {round(avg, 2)}/5\n📅 الشهر: {monthly['total']} سؤال, {monthly['counted']} محسوب", parse_mode="Markdown")
    except:
        await update.message.reply_text("⚠️ رقم غير صحيح")

# ==================== معالجة الرسائل ====================

async def handle_message(update: Update, context):
    sender = update.effective_user.id
    is_teacher = sender in teacher_ids
    
    # ===== المدرس يرد على سؤال =====
    if is_teacher and sender in teacher_state:
        qid = teacher_state[sender]["question_id"]
        student_id = teacher_state[sender]["student_id"]
        if qid not in questions or questions[qid]["status"] != "taken":
            del teacher_state[sender]
            update_teacher_load(sender, -1)
            return await update.message.reply_text("❌ هذه الجلسة انتهت")
        
        student_name = questions[qid]["student_name"]
        try:
            if update.message.text:
                await context.bot.send_message(student_id, f"📩 **رد من المدرس**\n\n👨‍🎓 للطالب: {student_name}\n\n{update.message.text}", parse_mode="Markdown")
                questions[qid]["reply"] = update.message.text
            elif update.message.photo:
                await context.bot.send_photo(student_id, update.message.photo[-1].file_id, caption=f"📩 رد من المدرس\n👨‍🎓 للطالب: {student_name}")
                questions[qid]["reply"] = "[صورة]"
            elif update.message.video:
                await context.bot.send_video(student_id, update.message.video.file_id, caption=f"📩 رد من المدرس\n👨‍🎓 للطالب: {student_name}")
                questions[qid]["reply"] = "[فيديو]"
            elif update.message.audio:
                await context.bot.send_audio(student_id, update.message.audio.file_id, caption=f"📩 رد من المدرس\n👨‍🎓 للطالب: {student_name}")
                questions[qid]["reply"] = "[صوت]"
            elif update.message.voice:
                await context.bot.send_voice(student_id, update.message.voice.file_id, caption=f"📩 رد من المدرس\n👨‍🎓 للطالب: {student_name}")
                questions[qid]["reply"] = "[فويس]"
            elif update.message.document:
                await context.bot.send_document(student_id, update.message.document.file_id, caption=f"📩 رد من المدرس\n👨‍🎓 للطالب: {student_name}")
                questions[qid]["reply"] = "[ملف]"
            else:
                return await update.message.reply_text("نوع غير مدعوم")
            
            questions[qid]["reply_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_questions()
            await update.message.reply_text("✅ تم إرسال الرد")
        except Exception as e:
            await update.message.reply_text(f"❌ فشل: {e}")
        return
    
    if is_teacher:
        return await update.message.reply_text("❌ لا توجد جلسة نشطة. استلم سؤال من الجروب")
    
    # ===== الطالب في جلسة نشطة - يرسل رسالة للمدرس مع أزرار =====
    if sender in student_active_session:
        teacher_id = student_active_session[sender]["teacher_id"]
        if teacher_id not in teacher_state:
            del student_active_session[sender]
            return await update.message.reply_text("❌ الجلسة انتهت")
        
        student_name = update.effective_user.full_name
        
        qid = f"{sender}_{datetime.now().timestamp()}"
        
        if update.message.text:
            text = update.message.text
            file_id = None
            file_type = None
            file_name = None
        elif update.message.photo:
            file_id = update.message.photo[-1].file_id
            file_type = "photo"
            file_name = None
            text = "📷 صورة"
        elif update.message.video:
            file_id = update.message.video.file_id
            file_type = "video"
            file_name = update.message.video.file_name
            text = "🎥 فيديو"
        elif update.message.audio:
            file_id = update.message.audio.file_id
            file_type = "audio"
            file_name = update.message.audio.file_name
            text = "🎵 صوت"
        elif update.message.voice:
            file_id = update.message.voice.file_id
            file_type = "voice"
            file_name = None
            text = "🎤 فويس"
        elif update.message.document:
            file_id = update.message.document.file_id
            file_type = "document"
            file_name = update.message.document.file_name or "ملف"
            text = f"📄 ملف: {file_name}"
        else:
            text = "ملف غير مدعوم"
            file_id = None
            file_type = None
            file_name = None
        
        questions[qid] = {
            "student_id": sender,
            "student_name": student_name,
            "teacher_id": teacher_id,
            "status": "taken",
            "count_for_stats": True,
            "rating": "",
            "text": text,
            "file_id": file_id,
            "file_type": file_type,
            "file_name": file_name,
            "reply": "",
            "reply_time": "",
            "group_message_id": None
        }
        save_questions()
        
        await send_question_to_teacher(context, teacher_id, qid, questions[qid])
        await update.message.reply_text("✅ تم إرسال سؤالك للمدرس مع الأزرار")
        return
    
    # ===== طالب جديد يرسل سؤال =====
    msg_id = update.message.message_id
    student_messages_in_group.setdefault(sender, []).append({"message_id": msg_id})
    
    qid = f"{sender}_{datetime.now().timestamp()}"
    student_name = update.effective_user.full_name
    
    if update.message.text:
        text, file_id, file_type, file_name = update.message.text, None, None, None
    elif update.message.photo:
        file_id, file_type, text = update.message.photo[-1].file_id, "photo", "📷 صورة"
        file_name = None
    elif update.message.video:
        file_id, file_type, text = update.message.video.file_id, "video", "🎥 فيديو"
        file_name = update.message.video.file_name
    elif update.message.audio:
        file_id, file_type, text = update.message.audio.file_id, "audio", "🎵 صوت"
        file_name = update.message.audio.file_name
    elif update.message.voice:
        file_id, file_type, text = update.message.voice.file_id, "voice", "🎤 فويس"
        file_name = None
    elif update.message.document:
        file_id, file_type = update.message.document.file_id, "document"
        file_name = update.message.document.file_name or "ملف"
        text = f"📄 ملف: {file_name}"
    else:
        text, file_id, file_type = "ملف غير مدعوم", None, None
        file_name = None
    
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("📥 استلام السؤال", callback_data=f"take|{qid}|{sender}")]])
    msg = f"💬 **سؤال جديد**\n👨‍🎓 {student_name}\n🔢 {sender}\n\n📝 {text}"
    
    try:
        sent_msg = None
        if file_id:
            if file_type == "photo":
                sent_msg = await context.bot.send_photo(GROUP_ID, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif file_type == "video":
                sent_msg = await context.bot.send_video(GROUP_ID, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif file_type == "audio":
                sent_msg = await context.bot.send_audio(GROUP_ID, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif file_type == "voice":
                sent_msg = await context.bot.send_voice(GROUP_ID, file_id, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
            elif file_type == "document":
                sent_msg = await context.bot.send_document(GROUP_ID, file_id, filename=file_name, caption=msg, reply_markup=keyboard, parse_mode="Markdown")
        else:
            sent_msg = await context.bot.send_message(GROUP_ID, msg, reply_markup=keyboard, parse_mode="Markdown")
        
        group_message_id = sent_msg.message_id if sent_msg else None
        
        questions[qid] = {
            "student_id": sender, "student_name": student_name, "teacher_id": None,
            "status": "open", "count_for_stats": True, "rating": "", "text": text,
            "file_id": file_id, "file_type": file_type, "file_name": file_name,
            "reply": "", "reply_time": "",
            "group_message_id": group_message_id
        }
        save_questions()
        
        await update.message.reply_text("✅ تم استلام سؤالك\n❌ /cancel للإلغاء")
    except Exception as e:
        await update.message.reply_text(f"❌ فشل: {e}")

# ==================== معالجة الأزرار ====================

async def button(update: Update, context):
    query = update.callback_query
    data = query.data
    teacher_id = query.from_user.id
    
    if data.startswith("cancel|"):
        qid = data.split("|")[1]
        if qid in questions and questions[qid]["status"] == "open":
            if questions[qid].get("group_message_id"):
                try:
                    await context.bot.delete_message(GROUP_ID, questions[qid]["group_message_id"])
                except:
                    pass
            del questions[qid]
            save_questions()
            if query.from_user.id in student_messages_in_group:
                for msg in student_messages_in_group[query.from_user.id]:
                    try:
                        await context.bot.delete_message(GROUP_ID, msg.get("message_id"))
                    except:
                        pass
                del student_messages_in_group[query.from_user.id]
            await query.answer("✅ تم الإلغاء")
            await query.edit_message_text("✅ تم إلغاء السؤال")
        return
    
    if data == "cancel_none":
        await query.answer("❌ تم الإلغاء")
        await query.edit_message_reply_markup()
        return
    
    if teacher_id not in teacher_ids:
        return await query.answer("❌ غير مصرح")
    
    # ===== استلام السؤال =====
    if data.startswith("take|"):
        if teacher_id in teacher_state:
            return await query.answer("❌ لديك سؤال نشط حالياً. قم بالرد عليه أو إنهائه أولاً.")
        
        current_load = teacher_active_questions.get(teacher_id, 0)
        if current_load >= MAX_ACTIVE_QUESTIONS:
            return await query.answer(f"❌ لديك سؤال نشط حالياً. لا يمكنك استلام سؤال جديد.")
        
        _, qid, student_id = data.split("|")
        student_id = int(student_id)
        if qid not in questions or questions[qid]["status"] != "open":
            return await query.answer("⚠️ السؤال غير متاح")
        if student_id in student_active_session:
            return await query.answer("⚠️ هذا الطالب لديه جلسة مع مدرس آخر")
        
        await query.answer("⏳ جاري...")
        
        # إزالة الأزرار من رسالة الجروب
        if questions[qid].get("group_message_id"):
            await remove_buttons_from_group_message(context, GROUP_ID, questions[qid]["group_message_id"])
        
        deleted = await delete_student_messages(context, student_id, GROUP_ID)
        sent = await send_all_questions_to_teacher(context, student_id, teacher_id)
        
        questions[qid]["teacher_id"] = teacher_id
        questions[qid]["status"] = "taken"
        teacher_state[teacher_id] = {"question_id": qid, "student_id": student_id}
        student_active_session[student_id] = {"teacher_id": teacher_id, "question_id": qid}
        update_teacher_load(teacher_id, 1)
        save_questions()
        
        await send_question_to_teacher(context, teacher_id, qid, questions[qid])
        
        await context.bot.send_message(GROUP_ID, f"✅ تم استلام أسئلة الطالب {questions[qid]['student_name']} بواسطة المدرس {query.from_user.full_name}")
    
    # ===== لا يتحسب =====
    elif data.startswith("skip|"):
        qid = data.split("|")[1]
        if qid in questions:
            questions[qid]["count_for_stats"] = False
            update_monthly_stats(teacher_id, False)
            save_questions()
            await query.answer("✅ لن يُحسب في الإحصائيات")
            await query.edit_message_reply_markup()
    
    # ===== إعادة للجروب =====
    elif data.startswith("return|"):
        qid = data.split("|")[1]
        if qid in questions and teacher_id in teacher_state:
            confirm_keyboard = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ نعم، أعد السؤال", callback_data=f"confirm_return|{qid}"),
                    InlineKeyboardButton("❌ إلغاء", callback_data=f"cancel_return|{qid}")
                ]
            ])
            await query.answer()
            await query.edit_message_reply_markup(reply_markup=confirm_keyboard)
    
    elif data.startswith("confirm_return|"):
        qid = data.split("|")[1]
        if qid in questions and teacher_id in teacher_state:
            await query.answer("🔄 جاري إعادة السؤال للجروب...")
            success = await return_to_group(context, teacher_id, qid)
            if success:
                await query.edit_message_reply_markup(reply_markup=None)
                await context.bot.send_message(teacher_id, "✅ تم إعادة السؤال للجروب بنجاح.")
            else:
                await query.answer("❌ فشل إعادة السؤال")
    
    elif data.startswith("cancel_return|"):
        qid = data.split("|")[1]
        await query.answer("❌ تم إلغاء الإعادة")
        if qid in questions:
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("🚫 لا يتحسب", callback_data=f"skip|{qid}")],
                [InlineKeyboardButton("🔄 إعادة للجروب", callback_data=f"return|{qid}")],
                [InlineKeyboardButton("🔚 إنهاء الجلسة", callback_data=f"end|{qid}")]
            ])
            await query.edit_message_reply_markup(reply_markup=keyboard)
    
    # ===== إنهاء الجلسة =====
    elif data.startswith("end|"):
        qid = data.split("|")[1]
        if teacher_id in teacher_state:
            student_id = teacher_state[teacher_id]["student_id"]
            teacher_name = query.from_user.full_name
            
            await send_end_session_message(context, student_id, teacher_name)
            
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("⭐1", f"rate|1|{teacher_id}"), InlineKeyboardButton("⭐⭐2", f"rate|2|{teacher_id}"), InlineKeyboardButton("⭐⭐⭐3", f"rate|3|{teacher_id}")],
                [InlineKeyboardButton("⭐⭐⭐⭐4", f"rate|4|{teacher_id}"), InlineKeyboardButton("⭐⭐⭐⭐⭐5", f"rate|5|{teacher_id}")]
            ])
            try:
                await context.bot.send_message(
                    student_id,
                    f"📝 **كيف تقيم رد المدرس {teacher_name}؟**\n\nاختر عدد النجوم:",
                    reply_markup=keyboard,
                    parse_mode="Markdown"
                )
                pending_ratings[student_id] = {"teacher_id": teacher_id, "question_id": qid}
            except Exception as e:
                logger.error(f"فشل إرسال طلب التقييم: {e}")
            
            if student_id in student_active_session:
                del student_active_session[student_id]
            del teacher_state[teacher_id]
            update_teacher_load(teacher_id, -1)
            
            if qid in questions:
                questions[qid]["status"] = "ended"
                save_questions()
            
            await query.answer("🔒 انتهت الجلسة")
            await query.edit_message_reply_markup()
            await context.bot.send_message(GROUP_ID, f"ℹ️ المدرس {teacher_name} أنهى الجلسة")
    
    # ===== تقييم =====
    elif data.startswith("rate|"):
        _, rating, teacher_id = data.split("|")
        rating = int(rating)
        teacher_id = int(teacher_id)
        teacher_ratings.setdefault(teacher_id, {"total_rating": 0, "count": 0})
        teacher_ratings[teacher_id]["total_rating"] += rating
        teacher_ratings[teacher_id]["count"] += 1
        update_monthly_stats(teacher_id, True, rating)
        
        if query.from_user.id in pending_ratings:
            qid = pending_ratings[query.from_user.id].get("question_id")
            if qid and qid in questions:
                questions[qid]["rating"] = rating
                save_questions()
            del pending_ratings[query.from_user.id]
        
        save_ratings()
        try:
            await context.bot.send_message(teacher_id, f"⭐ **تقييم جديد!**\n\nالطالب أعطاك {rating}/5 نجوم\n📊 متوسط تقييمك الحالي: {round(teacher_ratings[teacher_id]['total_rating'] / teacher_ratings[teacher_id]['count'], 2)}/5", parse_mode="Markdown")
        except:
            pass
        await query.answer(f"✅ {rating}/5")
        await query.edit_message_reply_markup()

# ==================== تشغيل البوت مع إعادة تشغيل تلقائي ====================

async def run_bot_with_retry():
    """تشغيل البوت مع إعادة تشغيل تلقائي عند الخطأ"""
    while True:
        try:
            logger.info("=" * 50)
            logger.info("🚀 تشغيل البوت...")
            logger.info("=" * 50)
            
            # تشغيل البوت
            await app.initialize()
            await app.start()
            await app.updater.start_polling(
                drop_pending_updates=True,
                poll_interval=1.0,
                timeout=30
            )
            
            # انتظر حتى يتوقف
            await asyncio.Event().wait()
            
        except asyncio.CancelledError:
            logger.info("⚠️ تم إلغاء التشغيل")
            break
            
        except Exception as e:
            logger.error(f"❌ خطأ في البوت: {e}")
            logger.info("🔄 جاري إعادة التشغيل بعد 10 ثواني...")
            
            # إيقاف البوت بشكل آمن
            try:
                await app.stop()
            except:
                pass
            
            await asyncio.sleep(10)
        
        finally:
            try:
                await app.stop()
            except:
                pass

# ==================== تحميل البيانات وتشغيل البوت ====================

load_teachers()
load_supervisors()
load_ratings()
load_questions()
load_monthly_stats()

app = ApplicationBuilder().token(TOKEN).build()

async def post_init(application):
    await check_admin_status(application.bot, GROUP_ID)

app.post_init = post_init

app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("help", help_command))
app.add_handler(CommandHandler("myid", myid))
app.add_handler(CommandHandler("mystats", mystats))
app.add_handler(CommandHandler("myrating", myrating))
app.add_handler(CommandHandler("mystatus", mystatus))
app.add_handler(CommandHandler("report", report))
app.add_handler(CommandHandler("cancel", cancel_question))
app.add_handler(CommandHandler("backup", backup_command))
app.add_handler(CommandHandler("startmonth", start_month))
app.add_handler(CommandHandler("closemonth", close_month))
app.add_handler(CommandHandler("monthstatus", month_status))
app.add_handler(CommandHandler("checkadmin", check_admin))
app.add_handler(CommandHandler("teachers", teachers_list))
app.add_handler(CommandHandler("addteacher", add_teacher))
app.add_handler(CommandHandler("removeteacher", remove_teacher))
app.add_handler(CommandHandler("addsupervisor", add_supervisor))
app.add_handler(CommandHandler("removesupervisor", remove_supervisor))
app.add_handler(CommandHandler("listsupervisors", list_supervisors))
app.add_handler(CommandHandler("allreplies", all_replies))
app.add_handler(CommandHandler("teacherstats", teacher_stats))

all_filter = (filters.TEXT & ~filters.COMMAND) | filters.PHOTO | filters.VIDEO | filters.AUDIO | filters.VOICE | filters.Document.ALL
app.add_handler(MessageHandler(all_filter, handle_message))
app.add_handler(CallbackQueryHandler(button))

print("=" * 50)
print("🤖 البوت شغال 🔥")
print(f"👨‍🏫 المدرسين: {len(teacher_ids)}")
print(f"👁️ المراقبين: {len(supervisor_ids)}")
print(f"📅 الشهر: {current_month} ({'مفتوح' if month_active else 'مغلق'})")
print(f"📌 كل مدرس: سؤال واحد فقط في المرة الواحدة")
print("=" * 50)

if __name__ == "__main__":
    try:
        asyncio.run(run_bot_with_retry())
    except KeyboardInterrupt:
        print("🛑 تم إيقاف البوت يدوياً")